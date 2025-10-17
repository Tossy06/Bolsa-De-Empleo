
# companies/views.py (o jobs/views.py)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Job
from .forms import JobForm

from django.http import HttpResponse
from datetime import datetime, timedelta
from companies.models import EmployeeQuotaTracking, QuotaHistoricalRecord, HiringReport

from django.http import JsonResponse
from .models import HiringReport, HiringReportStatus
from .forms import HiringReportForm

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from jobs.inclusive_language import InclusiveLanguageValidator

@login_required
def create_job_view(request):
    """Vista para crear una nueva oferta de trabajo"""
    if request.method == 'POST':
        form = JobForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.company = request.user
            job.save()
            
            # Registrar en auditoría
            from .models import AuditLog
            AuditLog.log_action(
                job=job,
                action='created',
                user=request.user,
                details={
                    'compliance_status': job.compliance_status,
                    'is_compliant': job.is_compliant(),
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                notes='Oferta creada y marcada como pendiente de revisión'
            )
            
            messages.success(
                request, 
                f'✅ Oferta "{job.title}" creada exitosamente. '
                'Está pendiente de revisión por un administrador antes de publicarse.'
            )
            return redirect('core:dashboard')
        else:
            # 🔥 MOSTRAR ERRORES ESPECÍFICOS
            messages.error(
                request, 
                '❌ Por favor corrige los errores en el formulario.'
            )
            
            # Mostrar cada error específico
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, f'Error: {error}')
                    else:
                        field_label = form.fields[field].label or field
                        messages.error(request, f'{field_label}: {error}')
    else:
        form = JobForm()

    context = {
        'form': form,
        'title': 'Crear nueva oferta de trabajo'
    }
    return render(request, 'companies/create_job.html', context)

@login_required
def edit_job_view(request, job_id):
    """Vista para editar una oferta existente"""
    job = get_object_or_404(Job, id=job_id, company=request.user)
    
    if request.method == 'POST':
        form = JobForm(request.POST, instance=job)
        if form.is_valid():
            old_status = job.compliance_status  #GUARDAR ESTADO ANTERIOR
            job = form.save()
            
            #Registrar actualización
            from .models import AuditLog
            AuditLog.log_action(
                job=job,
                action='updated',
                user=request.user,
                details={
                    'old_status': old_status,
                    'new_status': job.compliance_status,
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                notes='Oferta actualizada, requiere nueva revisión'
            )
            
            messages.success(request, f'Oferta "{job.title}" actualizada correctamente.')
            return redirect('core:dashboard')  # 🔧 Cambiar de 'companies:dashboard' a 'core:dashboard'
        else:
            messages.error(request, 'Corrige los errores en el formulario.')
    else:
        form = JobForm(instance=job)

    context = {
        'form': form,
        'job': job,
        'title': f'Editar: {job.title}'
    }
    return render(request, 'companies/create_job.html', context)

@login_required
def toggle_job_status(request, job_id):
    """Vista para activar/desactivar una oferta"""
    job = get_object_or_404(Job, id=job_id, company=request.user)
    job.is_active = not job.is_active
    job.save()
    
    status = "activada" if job.is_active else "desactivada"
    messages.success(request, f'Oferta "{job.title}" {status} correctamente.')
    
    return redirect('core:dashboard')

# Añadir esta función al final del archivo
@login_required
@require_http_methods(["POST"])
def validate_language_ajax(request):
    """Valida lenguaje en tiempo real"""
    text = request.POST.get('text', '')
    
    if not text:
        return JsonResponse({'valid': True, 'issues': []})
    
    issues = InclusiveLanguageValidator.scan_text(text)
    
    return JsonResponse({
        'valid': len(issues) == 0,
        'issues': issues
    })

@login_required
def hiring_reports_list_view(request):
    """
    Lista de informes de contratación de la empresa
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    reports = HiringReport.objects.filter(
        company=request.user
    ).select_related('job').order_by('-created_at')
    
    # Estadísticas
    stats = {
        'total': reports.count(),
        'pending': reports.filter(status=HiringReportStatus.PENDING).count(),
        'sent': reports.filter(status=HiringReportStatus.SENT).count(),
        'confirmed': reports.filter(status=HiringReportStatus.CONFIRMED).count(),
        'failed': reports.filter(status=HiringReportStatus.FAILED).count(),
    }
    
    context = {
        'reports': reports,
        'stats': stats,
    }
    
    return render(request, 'companies/hiring_reports_list.html', context)


@login_required
def create_hiring_report_view(request):
    """
    Crear nuevo informe de contratación
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        form = HiringReportForm(request.POST, user=request.user)
        if form.is_valid():
            report = form.save()
            messages.success(
                request,
                f'Informe de contratación #{report.contract_number} creado exitosamente. '
                'Ahora puede generar y enviar el informe al Ministerio.'
            )
            return redirect('company:hiring_reports')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = HiringReportForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Registrar Nueva Contratación',
    }
    
    return render(request, 'companies/hiring_report_form.html', context)


@login_required
def edit_hiring_report_view(request, report_id):
    """
    Editar informe de contratación (solo si no ha sido confirmado)
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    # No permitir editar si ya está confirmado
    if report.status == HiringReportStatus.CONFIRMED:
        messages.warning(
            request,
            'No puede editar un informe que ya fue confirmado por el Ministerio.'
        )
        return redirect('company:hiring_reports')
    
    if request.method == 'POST':
        form = HiringReportForm(request.POST, instance=report, user=request.user)
        if form.is_valid():
            report = form.save()
            messages.success(request, 'Informe actualizado exitosamente.')
            return redirect('company:hiring_reports')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = HiringReportForm(instance=report, user=request.user)
    
    context = {
        'form': form,
        'report': report,
        'title': f'Editar Informe #{report.contract_number}',
    }
    
    return render(request, 'companies/hiring_report_form.html', context)


@login_required
def hiring_report_detail_view(request, report_id):
    """
    Ver detalles de un informe específico
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    report = get_object_or_404(
        HiringReport,
        id=report_id,
        company=request.user
    )
    
    context = {
        'report': report,
    }
    
    return render(request, 'companies/hiring_report_detail.html', context)


@login_required
def delete_hiring_report_view(request, report_id):
    """
    Eliminar informe (solo si no ha sido enviado)
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    # Solo permitir eliminar si está pendiente
    if report.status not in [HiringReportStatus.PENDING, HiringReportStatus.FAILED]:
        messages.warning(
            request,
            'No puede eliminar un informe que ya fue enviado al Ministerio.'
        )
        return redirect('company:hiring_reports')
    
    if request.method == 'POST':
        contract_number = report.contract_number
        report.delete()
        messages.success(
            request,
            f'Informe #{contract_number} eliminado exitosamente.'
        )
        return redirect('company:hiring_reports')
    
    context = {
        'report': report,
    }
    
    return render(request, 'companies/hiring_report_confirm_delete.html', context)


@login_required
def send_hiring_report_view(request, report_id):
    """
    Enviar informe al Ministerio (genera PDF/XML y simula envío)
    """
    if request.user.user_type != 'company':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    # Verificar que esté pendiente
    if report.status not in [HiringReportStatus.PENDING, HiringReportStatus.FAILED, HiringReportStatus.RETRY]:
        return JsonResponse({
            'success': False,
            'error': 'Este informe ya fue enviado o está en proceso'
        }, status=400)
    
    try:
        # Aquí llamaremos a la función que genera el PDF y envía
        # Por ahora, lo dejamos preparado
        from .services.ministry_report_service import generate_and_send_report
        
        result = generate_and_send_report(report)
        
        if result['success']:
            messages.success(
                request,
                f'Informe enviado exitosamente al Ministerio. '
                f'Número de recibo: {result.get("receipt_number", "N/A")}'
            )
            return JsonResponse({'success': True, 'data': result})
        else:
            messages.error(
                request,
                f'Error al enviar el informe: {result.get("error", "Error desconocido")}'
            )
            return JsonResponse({'success': False, 'error': result.get('error')}, status=500)
    
    except Exception as e:
        messages.error(request, f'Error al procesar el informe: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required
def download_my_report_pdf(request, report_id):
    """
    Genera y descarga el PDF en tiempo real (sin guardarlo en servidor)
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Acceso denegado.')
        return redirect('core:dashboard')
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    try:
        # Generar PDF en memoria
        pdf_content = generate_pdf_report(report)
        
        # Retornar como descarga directa
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="informe_{report.contract_number}.pdf"'
        
        return response
    
    except Exception as e:
        messages.error(request, f'Error al generar el PDF: {str(e)}')
        return redirect('company:hiring_report_detail', report_id=report_id)


@login_required
def download_my_report_xml(request, report_id):
    """
    Genera y descarga el XML en tiempo real (sin guardarlo en servidor)
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Acceso denegado.')
        return redirect('core:dashboard')
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    try:
        # Generar XML en memoria
        xml_content = generate_xml_report(report)
        
        # Retornar como descarga directa
        response = HttpResponse(xml_content, content_type='application/xml')
        response['Content-Disposition'] = f'attachment; filename="informe_{report.contract_number}.xml"'
        
        return response
    
    except Exception as e:
        messages.error(request, f'Error al generar el XML: {str(e)}')
        return redirect('company:hiring_report_detail', report_id=report_id)


# TAMBIÉN ACTUALIZAR LA VISTA DE CREAR (sin generar archivos)
@login_required
def create_hiring_report_view(request):
    """
    Crear nuevo informe de contratación (SIN generar PDF/XML)
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        form = HiringReportForm(request.POST, user=request.user)
        if form.is_valid():
            report = form.save()
            
            # Generar firma digital (sin archivos)
            report.generate_signature()
            
            messages.success(
                request,
                f'✅ Informe #{report.contract_number} creado exitosamente. '
                'Ahora puede descargarlo en formato PDF o XML.'
            )
            
            return redirect('company:hiring_report_detail', report_id=report.id)
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = HiringReportForm(user=request.user)
    
    context = {
        'form': form,
        'title': 'Registrar Nueva Contratación',
    }
    
    return render(request, 'companies/hiring_report_form.html', context)


# VISTA PARA SIMULAR ENVÍO AL MINISTERIO
@login_required
def send_hiring_report_view(request, report_id):
    """
    Simular envío al Ministerio (sin guardar PDF/XML)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    if request.user.user_type != 'company':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)
    
    report = get_object_or_404(HiringReport, id=report_id, company=request.user)
    
    # Verificar que esté pendiente
    if report.status not in [HiringReportStatus.PENDING, HiringReportStatus.FAILED, HiringReportStatus.RETRY]:
        return JsonResponse({
            'success': False,
            'error': 'Este informe ya fue enviado'
        }, status=400)
    
    try:
        from companies.services.ministry_report_service import send_to_ministry_api
        import hashlib
        from datetime import datetime
        
        # Generar firma si no existe
        if not report.digital_signature:
            report.generate_signature()
        
        # Simular envío
        result = send_to_ministry_api(report)
        
        if result['success']:
            # Marcar como confirmado
            report.mark_as_confirmed(
                receipt_number=result['receipt_number'],
                response_data=result.get('response_data', {})
            )
            
            messages.success(
                request,
                f'✅ Informe enviado exitosamente. Número de recibo: {result["receipt_number"]}'
            )
            return JsonResponse({'success': True, 'data': result})
        else:
            # Manejar fallo
            if report.can_retry():
                report.increment_retry()
                error_msg = f'Error en el envío. Intento {report.retry_count} de 3.'
            else:
                report.mark_as_failed(result.get('error', 'Error desconocido'))
                error_msg = 'Se agotaron los intentos de envío.'
            
            messages.error(request, f'❌ {error_msg}')
            return JsonResponse({'success': False, 'error': error_msg}, status=500)
    
    except Exception as e:
        report.mark_as_failed(str(e))
        messages.error(request, f'Error: {str(e)}')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
@login_required
def quota_compliance_dashboard(request):
    """
    Dashboard principal de cumplimiento de cuotas
    """
    if request.user.user_type != 'company':
        messages.error(request, 'Esta sección es solo para empresas.')
        return redirect('core:dashboard')
    
    # Obtener o crear tracking de cuotas
    tracking, created = EmployeeQuotaTracking.objects.get_or_create(
        company=request.user
    )
    
    # Si se acaba de crear, actualizar desde reportes
    if created:
        tracking.update_from_reports()
    
    # Obtener histórico (últimos 12 meses)
    historical_data = QuotaHistoricalRecord.objects.filter(
        company=request.user
    ).order_by('-period')[:12]
    
    # Filtro de periodo
    period_filter = request.GET.get('period', 'monthly')
    
    # Datos para gráfico
    chart_labels = []
    chart_data_compliance = []
    chart_data_required = []
    chart_data_current = []
    
    for record in reversed(list(historical_data)):
        chart_labels.append(record.period.strftime('%b %Y'))
        chart_data_compliance.append(float(record.compliance_percentage))
        chart_data_required.append(record.required_employees)
        chart_data_current.append(record.employees_with_disability)
    
    # Informes recientes
    recent_reports = HiringReport.objects.filter(
        company=request.user
    ).order_by('-created_at')[:5]
    
    context = {
        'tracking': tracking,
        'historical_data': historical_data,
        'chart_labels': chart_labels,
        'chart_data_compliance': chart_data_compliance,
        'chart_data_required': chart_data_required,
        'chart_data_current': chart_data_current,
        'period_filter': period_filter,
        'recent_reports': recent_reports,
    }
    
    return render(request, 'companies/quota_compliance_dashboard.html', context)


@login_required
def update_employee_count(request):
    """
    Vista para actualizar el conteo de empleados
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    if request.user.user_type != 'company':
        return JsonResponse({'success': False, 'error': 'Acceso denegado'}, status=403)
    
    try:
        total_employees = int(request.POST.get('total_employees', 0))
        
        if total_employees < 0:
            return JsonResponse({
                'success': False,
                'error': 'El número de empleados no puede ser negativo'
            }, status=400)
        
        tracking, _ = EmployeeQuotaTracking.objects.get_or_create(
            company=request.user
        )
        
        tracking.total_employees = total_employees
        tracking.update_from_reports()  # Actualiza empleados con discapacidad
        tracking.save()
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_employees': tracking.total_employees,
                'employees_with_disability': tracking.employees_with_disability,
                'required_employees': tracking.required_employees_with_disability,
                'compliance_percentage': tracking.compliance_percentage,
                'is_compliant': tracking.is_compliant,
                'employees_needed': tracking.employees_needed
            }
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def export_quota_report_pdf(request):
    """
    Exportar informe de cuotas en PDF
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    from io import BytesIO
    
    if request.user.user_type != 'company':
        messages.error(request, 'Acceso denegado.')
        return redirect('core:dashboard')
    
    tracking = EmployeeQuotaTracking.objects.get(company=request.user)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    elements.append(Paragraph(
        "INFORME DE CUMPLIMIENTO DE CUOTAS DE EMPLEO",
        styles['Title']
    ))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información de la empresa
    elements.append(Paragraph(f"<b>Empresa:</b> {request.user.get_full_name()}", styles['Normal']))
    elements.append(Paragraph(f"<b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Datos de cumplimiento
    data = [
        ['Concepto', 'Valor'],
        ['Total de Empleados', str(tracking.total_employees)],
        ['Empleados con Discapacidad', str(tracking.employees_with_disability)],
        ['Empleados Requeridos (2%)', str(tracking.required_employees_with_disability)],
        ['Porcentaje de Cumplimiento', f"{tracking.compliance_percentage}%"],
        ['Estado', 'CUMPLE' if tracking.is_compliant else 'NO CUMPLE'],
    ]
    
    if not tracking.is_compliant:
        data.append(['Empleados Faltantes', str(tracking.employees_needed)])
    
    table = Table(data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cumplimiento_cuotas_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    return response


@login_required
def export_quota_report_excel(request):
    """
    Exportar informe de cuotas en Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    if request.user.user_type != 'company':
        messages.error(request, 'Acceso denegado.')
        return redirect('core:dashboard')
    
    tracking = EmployeeQuotaTracking.objects.get(company=request.user)
    historical = QuotaHistoricalRecord.objects.filter(
        company=request.user
    ).order_by('-period')
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cumplimiento de Cuotas"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Título
    ws['A1'] = 'INFORME DE CUMPLIMIENTO DE CUOTAS DE EMPLEO'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Información general
    ws['A3'] = 'Empresa:'
    ws['B3'] = request.user.get_full_name()
    ws['A4'] = 'Fecha:'
    ws['B4'] = datetime.now().strftime('%d/%m/%Y')
    
    # Datos actuales
    ws['A6'] = 'SITUACIÓN ACTUAL'
    ws['A6'].font = Font(bold=True)
    ws.merge_cells('A6:B6')
    
    current_data = [
        ['Total de Empleados', tracking.total_employees],
        ['Empleados con Discapacidad', tracking.employees_with_disability],
        ['Empleados Requeridos (2%)', tracking.required_employees_with_disability],
        ['Porcentaje de Cumplimiento', f"{tracking.compliance_percentage}%"],
        ['Estado', 'CUMPLE' if tracking.is_compliant else 'NO CUMPLE'],
    ]
    
    row = 7
    for label, value in current_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Histórico
    ws[f'A{row+1}'] = 'HISTÓRICO MENSUAL'
    ws[f'A{row+1}'].font = Font(bold=True)
    ws.merge_cells(f'A{row+1}:E{row+1}')
    
    headers = ['Periodo', 'Total Empleados', 'Empleados c/Discapacidad', 'Requeridos', 'Cumplimiento %']
    header_row = row + 2
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    data_row = header_row + 1
    for record in historical:
        ws.cell(row=data_row, column=1, value=record.period.strftime('%m/%Y'))
        ws.cell(row=data_row, column=2, value=record.total_employees)
        ws.cell(row=data_row, column=3, value=record.employees_with_disability)
        ws.cell(row=data_row, column=4, value=record.required_employees)
        ws.cell(row=data_row, column=5, value=f"{record.compliance_percentage}%")
        data_row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cumplimiento_cuotas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response